import re
import copy
from openpyxl.utils import get_column_letter
from openpyxl.cell import MergedCell

def insert_rows(self, row_idx, cnt, above=False, copy_style=True, fill_formulae=True):
    """
    Inserts new (empty) rows into worksheet at specified row index.
    Fully compatible with openpyxl >= 3.1.5.
    """

    CELL_RE = re.compile(r"(?P<col>\$?[A-Z]+)(?P<row>\$?\d+)")

    if not hasattr(self, "formula_attributes"):
        self.formula_attributes = {}

    row_idx = row_idx - 1 if above else row_idx

    def replace(m):
        row = m.group("row")
        prefix = "$" if "$" in row else ""
        row = int(row.replace("$", ""))
        row += cnt if row > row_idx else 0
        return m.group("col") + prefix + str(row)

    # --- Shift existing cells down ---
    old_cells = set()
    old_fas = set()
    new_cells = {}
    new_fas = {}

    for c in list(self._cells.values()):
        if isinstance(c, MergedCell):
            continue  # skip merged placeholders

        old_coor = c.coordinate

        # Update formulas
        if c.data_type == "f":
            c.value = CELL_RE.sub(replace, c.value)
            if (
                old_coor in self.formula_attributes
                and "ref" in self.formula_attributes[old_coor]
            ):
                self.formula_attributes[old_coor]["ref"] = CELL_RE.sub(
                    replace, self.formula_attributes[old_coor]["ref"]
                )

        # Shift row index
        if c.row > row_idx:
            old_cells.add((c.row, c.col_idx))
            c.row += cnt
            new_cells[(c.row, c.col_idx)] = c
            if old_coor in self.formula_attributes:
                old_fas.add(old_coor)
                fa = self.formula_attributes[old_coor].copy()
                new_fas[c.coordinate] = fa

    for coor in old_cells:
        self._cells.pop(coor, None)
    self._cells.update(new_cells)

    for fa in old_fas:
        self.formula_attributes.pop(fa, None)
    self.formula_attributes.update(new_fas)

    # --- Shift row dimensions ---
    if self.row_dimensions:
        for row in range(max(self.row_dimensions.keys()) + cnt, row_idx + cnt, -1):
            if row - cnt in self.row_dimensions:
                new_rd = copy.copy(self.row_dimensions[row - cnt])
                new_rd.index = row
                self.row_dimensions[row] = new_rd
                self.row_dimensions.pop(row - cnt, None)

    # --- Create new rows ---
    row_idx += 1
    for row in range(row_idx, row_idx + cnt):
        prev_row = self.row_dimensions.get(row - 1)
        if prev_row:
            new_rd = copy.copy(prev_row)
            new_rd.index = row
            self.row_dimensions[row] = new_rd

        for col in range(1, self.max_column + 1):
            col_letter = get_column_letter(col)
            cell = self.cell(row=row, column=col)
            source = self.cell(row=row - 1, column=col)

            if copy_style:
                cell.number_format = source.number_format
                cell.font = source.font.copy()
                cell.alignment = source.alignment.copy()
                cell.border = source.border.copy()
                cell.fill = source.fill.copy()

            if fill_formulae and source.data_type == "f":
                s_coor = source.coordinate
                if (
                    s_coor in self.formula_attributes
                    and "ref" not in self.formula_attributes[s_coor]
                ):
                    fa = self.formula_attributes[s_coor].copy()
                    self.formula_attributes[cell.coordinate] = fa

                cell.value = re.sub(
                    rf"(\$?[A-Z]{{1,3}}\$?){row - 1}",
                    lambda m: m.group(1) + str(row),
                    source.value,
                )
                cell.data_type = "f"

    # --- Update merged cell ranges ---
    if hasattr(self, "merged_cells"):
        new_ranges = []
        for cr in self.merged_cells.ranges:
            new_ref = CELL_RE.sub(replace, str(cr))
            new_ranges.append(new_ref)
        self.merged_cells.ranges = new_ranges

# Monkey-patch onto Worksheet
from openpyxl.worksheet.worksheet import Worksheet
Worksheet.insert_rows = insert_rows

