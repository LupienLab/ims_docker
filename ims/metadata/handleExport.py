'''
Created on Dec. 8, 2020

@author: ankita
'''
from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponse
from django.http.response import HttpResponseRedirect, JsonResponse
from metadata.models import *
from metadata.forms import *
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Color, Fill
from openpyxl.utils.cell import get_column_letter
from datetime import datetime
from datetime import date
from metadata.excelRow import insert_rows
import json
import csv

now = datetime.now()
today = date.today()

ROOTFOLDER="/Users/ankita/Documents/eclipse-workspace/imsDB/ims_docker/ims"
#ROOTFOLDER="/Users/ankita/Documents/eclipse-workspace/imsDB/ims_docker/ims/"
dt_string = now.strftime("%Y-%m-%d_%H-%M-%S")
dt_today = today.strftime("%Y-%m-%d")


def handle_exportSequencingform(request,prj_pk):
    experimentList=request.POST.getlist('choose_experiments')
    buffer=request.POST.get('buffer')
    instructions=request.POST.get('instructions')
    data_recipient_contact_name=request.POST.get('data_recipient_contact_name')
    data_recipient_contact_email=request.POST.get('data_recipient_contact_email')
    grant=request.POST.get('grant')
    bp_length=request.POST.get('bp_length')
    low_diversity_sample=request.POST.get('low_diversity_sample')
    sequencing_type=request.POST.get('sequencing_type')
    multiplexing_sequencing=request.POST.get('multiplexing_sequencing')

    
    dt_string = now.strftime("%Y-%m-%d_%H-%M-%S")
    dt_today = today.strftime("%Y-%m-%d")
    project_id=prj_pk
    prj = Project.objects.get(pk=project_id)
      
    # Create the HttpResponse object with the appropriate CSV header.
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = "attachment; filename=Epigenetics_Initiative_Sample_Submission_Form_"+dt_string+".xlsx"
    file_path_new = 'static/Epigenetics_Initiative_Sample_Submission_Form_v3.1.xlsx'
 
    wb = load_workbook(file_path_new)
    ws = wb.worksheets[0]
    ws.insert_rows = insert_rows
    
    contributor1 = prj.created_by
    contributor2 = prj.contributor.all().order_by('first_name')
    
    membersList = []
    membersList.append(contributor1.get_full_name())
    for values in contributor2:
        membersList.append(values.get_full_name())
    m=list(set(membersList))
    
    emailList=[]
    emailList.append(contributor1.email)
    for values in contributor2:
        emailList.append(values.email)
    e=list(set(emailList))
    
    
    ws.cell(row=5, column=5).value = dt_today
    ws.cell(row=7, column=2).value = " / ".join(m)
    ws.cell(row=8, column=2).value = " / ".join(e)
    ws.cell(row=9, column=2).value = data_recipient_contact_name
    ws.cell(row=10, column=2).value = data_recipient_contact_email
    ws.cell(row=14, column=2).value = grant
    
    sampleRowNo=18
    count=0
    for ep in experimentList:
        #insert_rows(ws, sampleRowNo, 1, above=True, copy_style=True)
        insert_rows(ws, row_idx= sampleRowNo, cnt = 1, above=False, copy_style=True)
        exp=Experiment.objects.get(pk=int(ep))
        #ws.insert_rows(sampleRowNo, amount=1,)
        ws.cell(row=sampleRowNo, column=1).value = exp.uid
        ws.cell(row=sampleRowNo, column=2).value = exp.name
        ws.cell(row=sampleRowNo, column=3).value = exp.json_type.name
        
        ws.cell(row=sampleRowNo, column=4).value = exp.biosample.biosource.source_organism.name
        
#         if(exp.biosample.sample_type):
#             ws.cell(row=sampleRowNo, column=5).value = exp.biosample.sample_type.name
            
        if(exp.biosample_quantity_units):
            ws.cell(row=sampleRowNo, column=6).value = str(exp.biosample_quantity)+" "+str(exp.biosample_quantity_units.name)
        elif(exp.biosample_quantity):
            ws.cell(row=sampleRowNo, column=6).value = str(exp.biosample_quantity)
        else:
            ws.cell(row=sampleRowNo, column=6).value = "value not filled in db"
        
        ws.cell(row=sampleRowNo, column=12).value = buffer

        
        sampleRowNo += 1
        count += 1
    
    ws.cell(row=23+count, column=2).value = bp_length
    ws.cell(row=24+count, column=2).value = low_diversity_sample
    ws.cell(row=25+count, column=2).value = sequencing_type
    ws.cell(row=26+count, column=2).value = multiplexing_sequencing
    ws.cell(row=28+count, column=1).value = instructions
    
    
    wb.save(response)    
    return response



def exportSequencingform(request,prj_pk):
    if request.method == 'POST':
        form = SequencingForm(request.POST,initial={'prj_pk':prj_pk})
        if form.is_valid():
            res= handle_exportSequencingform(request,prj_pk)
            return res

    else:
        form = SequencingForm(initial={'prj_pk':prj_pk})
     
    pageContext = {
        'form': form,
        'form_name':'Export genomic core sequencing form'
        }
    return render(request, 'export.html', pageContext)

def handle_exportegaform(request,prj_pk):
    experimentList=request.POST.getlist('choose_experiments')
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = "attachment; filename=EGA_Sample_Submission_Form_"+dt_string+".xlsx"
    file_path_new = 'static/EGA_Array_based_Format_V4.3.xlsx'
 
    wb = load_workbook(file_path_new)
    ws = wb.worksheets[1]
    c=5
    sampleCount=0
    biosampleList=[]
    for exp in experimentList:
        biosample = Experiment.objects.get(pk=exp).biosample
        if(biosample.name not in biosampleList):
            biosampleList.append(biosample.name)
            ws.cell(row=c, column=1).value = biosample.biosource.name
            ws.cell(row=c, column=2).value = biosample.name
            ws.cell(row=c, column=3).value = biosample.description
            ws.cell(row=c, column=4).value = biosample.biosource.source_organism.name
            if("sex" in json.loads(biosample.biosource.json_fields)):
                ws.cell(row=c, column=5).value = json.loads(biosample.biosource.json_fields)["sex"]
            ws.cell(row=c, column=6).value = biosample.biosource.disease
            if("donor_id" in json.loads(biosample.biosource.json_fields)):
                ws.cell(row=c, column=7).value = json.loads(biosample.biosource.json_fields)["donor_id"]
            
            sampleCount+=1
            c+=1
    
    
    c=15
    ws = wb.worksheets[2]
    ws.cell(row=c, column=1).value = Project.objects.get(pk=prj_pk).name
    ws.cell(row=c, column=7).value = sampleCount
    
    c=13
    ws = wb.worksheets[3]
    for exp in experimentList:
        ex=Experiment.objects.get(pk=exp) 
        biosample = ex.biosample
        files = SeqencingFile.objects.filter(experiment=exp)
        for f in files:
            ws.cell(row=c, column=1).value = biosample.name
            if(f.run.sequencing_instrument):
                ws.cell(row=c, column=2).value = f.run.sequencing_instrument.name
            ws.cell(row=c, column=3).value = f.cluster_path
            ws.cell(row=c, column=4).value = f.file_format.name
            c+=1
    
    wb.save(response)    
    return response


def handle_exportiheacform(request,prj_pk):
    experimentList=request.POST.getlist('choose_experiments')
    response = HttpResponse(content_type='text/csv')
    
    typeList=[]
    sampleList=[]
    for exp in experimentList:
        biosample = Experiment.objects.get(pk=exp).biosample
        typeB = biosample.biosource.json_type.name
        typeList.append(typeB)
        
        if(biosample.json_type.name=="culture details - Available"):
            sampleList.append("Y")
        else:
            sampleList.append("N")
            
    c=0
    if(len(set(typeList)) == 1):
        if(typeList[0]=="cell line"):
            headerLine=["id", "sample_ontology_uri", "disease", "disease_ontology_uri", "biomaterial_type", "line", "lineage", 
                        "differentiation_stage", "medium", "sex"]
            response['Content-Disposition'] = "attachment; filename=ihec_cellLine_"+dt_string+".csv"
            writer = csv.writer(response)
            
            writer.writerow(headerLine)
                
            biosampleList=[]
            for exp in experimentList:
                ex=Experiment.objects.get(pk=exp) 
                biosample = ex.biosample
                if(biosample.name not in biosampleList):
                    biosampleList.append(biosample.name)
                    fields=[]
                    fields.append(biosample.name)
                    fields.append("")
                    fields.append(biosample.biosource.disease)
                    fields.append("")
                    fields.append(biosample.biosource.json_type.name)
                    fields.append(json.loads(biosample.biosource.json_fields)["cell_line_name"])
                    fields.append(json.loads(biosample.biosource.json_fields)["lineage"])
                    fields.append(json.loads(biosample.biosource.json_fields)["differentiation_stage"])
                    fields.append(json.loads(biosample.biosource.json_fields)["medium"])
                    fields.append(json.loads(biosample.biosource.json_fields)["sex"])
                    writer.writerow(fields)
                    c+=1
         
        if(typeList[0]=="tissue") or (typeList[0]=="tissue explants"):
            
            headerLine=["id", "sample_ontology_uri", "disease", "disease_ontology_uri", "biomaterial_type", "tissue_type", "tissue_depot", 
                        "donor_id", "donor_age", "donor_age_unit", "donor_health_status", "donor_sex", "donor_ethnicity"]
            response['Content-Disposition'] = "attachment; filename=ihec_primaryTissue_"+dt_string+".csv"
            writer = csv.writer(response)
            
            writer.writerow(headerLine)
            
            
            biosampleList=[]
            for exp in experimentList:
                ex=Experiment.objects.get(pk=exp) 
                biosample = ex.biosample
                if(biosample.name not in biosampleList):
                    biosampleList.append(biosample.name)
                    fields=[]
                    fields.append(biosample.name)
                    fields.append("")
                    fields.append(biosample.biosource.disease)
                    fields.append("")
                    fields.append(biosample.biosource.json_type.name)
                    fields.append(json.loads(biosample.biosource.json_fields)["tissue_type"])
                    fields.append("")
                    fields.append(json.loads(biosample.biosource.json_fields)["donor_id"])
                    fields.append(json.loads(biosample.biosource.json_fields)["donor_age"])
                    fields.append(json.loads(biosample.biosource.json_fields)["donor_age_unit"])
                    fields.append(json.loads(biosample.biosource.json_fields)["donor_health_status"])
                    fields.append(json.loads(biosample.biosource.json_fields)["sex"])
                    fields.append(json.loads(biosample.biosource.json_fields)["donor_ethnicity"])
                    writer.writerow(fields)
                    c+=1
         
        if(typeList[0]=="primary cell"):
            if(len(set(typeList)) == 1) and typeList[0]=="Y":
                headerLine=["id", "sample_ontology_uri", "disease", "disease_ontology_uri", "biomaterial_type", "cell_type", 
                            "donor_id", "donor_age", "donor_age_unit", "donor_health_status", "donor_sex", "donor_ethnicity"]
                response['Content-Disposition'] = "attachment; filename=ihec_primaryCell_"+dt_string+".csv"
                writer = csv.writer(response)
            
                writer.writerow(headerLine)
                biosampleList=[]
                for exp in experimentList:
                    ex=Experiment.objects.get(pk=exp) 
                    biosample = ex.biosample
                    if(biosample.name not in biosampleList):
                        biosampleList.append(biosample.name)
                        fields=[]
                        fields.append(biosample.name)
                        fields.append("")
                        fields.append(biosample.biosource.disease)
                        fields.append("")
                        fields.append(biosample.biosource.json_type.name)
                        fields.append(json.loads(biosample.biosource.json_fields)["cell_type"])
                        fields.append(json.loads(biosample.biosource.json_fields)["donor_id"])
                        fields.append(json.loads(biosample.biosource.json_fields)["donor_age"])
                        fields.append(json.loads(biosample.biosource.json_fields)["donor_age_unit"])
                        fields.append(json.loads(biosample.biosource.json_fields)["donor_health_status"])
                        fields.append(json.loads(biosample.biosource.json_fields)["sex"])
                        fields.append(json.loads(biosample.biosource.json_fields)["donor_ethnicity"])
                        writer.writerow(fields)
                        c+=1       
            else:
                headerLine=["id", "sample_ontology_uri", "disease", "disease_ontology_uri", "biomaterial_type", "cell_type", "culture_conditions", 
                            "donor_id", "donor_age", "donor_age_unit", "donor_health_status", "donor_sex", "donor_ethnicity"]
                response['Content-Disposition'] = "attachment; filename=ihec_primaryCellCulture_"+dt_string+".csv"
                writer = csv.writer(response)
            
                writer.writerow(headerLine)
                biosampleList=[]
                for exp in experimentList:
                    ex=Experiment.objects.get(pk=exp) 
                    biosample = ex.biosample
                    if(biosample.name not in biosampleList):
                        biosampleList.append(biosample.name)
                        fields=[]
                        fields.append(biosample.name)
                        fields.append("")
                        fields.append(biosample.biosource.disease)
                        fields.append("")
                        fields.append(biosample.biosource.json_type.name)
                        fields.append(json.loads(biosample.biosource.json_fields)["cell_type"])
                        fields.append("")
                        fields.append(json.loads(biosample.biosource.json_fields)["donor_id"])
                        fields.append(json.loads(biosample.biosource.json_fields)["donor_age"])
                        fields.append(json.loads(biosample.biosource.json_fields)["donor_age_unit"])
                        fields.append(json.loads(biosample.biosource.json_fields)["donor_health_status"])
                        fields.append(json.loads(biosample.biosource.json_fields)["sex"])
                        fields.append(json.loads(biosample.biosource.json_fields)["donor_ethnicity"])
                        writer.writerow(fields)
        
                        c+=1       
       
    else:
        pass
    
  
    return response


def handle_exportgeoform(request,prj_pk):
    experimentList=request.POST.getlist('choose_experiments')
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = "attachment; filename=GEO_sheet_v04282020_"+dt_string+".xlsx"
    file_path_new = 'static/GEO_sheet_v04282020.xlsx'
 
    wb = load_workbook(file_path_new)
    ws = wb.worksheets[0]
    ws.insert_rows = insert_rows
    
    ws.cell(row=9, column=2).value = Project.objects.get(pk=prj_pk).name
    ws.cell(row=10, column=2).value = Project.objects.get(pk=prj_pk).description
    ws.cell(row=12, column=2).value = Project.objects.get(pk=prj_pk).created_by.get_full_name()
    c=13
   
    for n in Project.objects.get(pk=prj_pk).contributor.all():
        if(c>14):
            insert_rows(ws, row_idx= c, cnt = 1, above=False, copy_style=True)
            
            ws.cell(row=c, column=1).value = "contributor"
        
        ws.cell(row=c, column=2).value = n.get_full_name()
        c+=1
    
    rowNo = ws.max_row
    for i in range (rowNo):
        if((ws.cell(row=i+1, column=1).value)=="Sample name"):
            desiredrowNo = i+2
            break
        
    c=desiredrowNo
    sampleNo=1
    biosampleList=[]
    for exp in experimentList:
        e=Experiment.objects.get(pk=exp) 
        biosample = e.biosample
        if(biosample.name not in biosampleList):
            biosampleList.append(biosample.name)
            insert_rows(ws, row_idx= c, cnt = 1, above=False, copy_style=True)
            files = SeqencingFile.objects.filter(experiment=e)
            ws.cell(row=c, column=1).value = biosample.name
            ws.cell(row=c, column=2).value = e.name
            ws.cell(row=c, column=3).value = biosample.biosource.name
            ws.cell(row=c, column=4).value = biosample.biosource.source_organism.name
            if (str(e.description)!="nan"):
                ws.cell(row=c, column=9).value = e.description
            colno=11
            for f in files:
                ws.cell(row=c, column=colno).value = f.name
                colno+=1
            c+=1
        
    rowNo = ws.max_row
    for i in range (rowNo):
        if((ws.cell(row=i+1, column=1).value)=="RAW FILES"):
            desiredrowNo = i+3
            break
        
    c=desiredrowNo
    
    for exp in experimentList:
        e=Experiment.objects.get(pk=exp)
        files = SeqencingFile.objects.filter(experiment=e)
        for f in files:
            insert_rows(ws, row_idx= c, cnt = 1, above=False, copy_style=False)
            ws.cell(row=c, column=1).value = f.name
            ws.cell(row=c, column=2).value = f.file_format.name
            ws.cell(row=c, column=3).value = f.md5sum
            if(f.run.sequencing_instrument):
                ws.cell(row=c, column=4).value = f.run.sequencing_instrument.name
            if(f.related_files):
                ws.cell(row=c, column=5).value = "paired-end"
            elif(f.paired_end=="index"):
                ws.cell(row=c, column=5).value = "index"
            else:
                ws.cell(row=c, column=5).value = "single"
            ws.cell(row=c, column=6).value = f.cluster_path
            c+=1
            
    rowNo = ws.max_row
    for i in range (rowNo):
        if((ws.cell(row=i+1, column=1).value)=="PAIRED-END EXPERIMENTS"):
            desiredrowNo = i+3
            break
        
    c=desiredrowNo
    
    listfiles=[]
    for exp in experimentList:
        e=Experiment.objects.get(pk=exp)
        files = SeqencingFile.objects.filter(experiment=e)
        for f in files:
            if((f.name not in listfiles) and (f.related_files)):
                insert_rows(ws, row_idx= c, cnt = 1, above=False, copy_style=False)
                ws.cell(row=c, column=1).value = f.name
                ws.cell(row=c, column=2).value = f.related_files.name
                listfiles.append(f.name)
                listfiles.append(f.related_files.name)
                c+=1
            
    wb.save(response)    
    return response



def exportform(request,prj_pk,slug):
    if request.method == 'POST':
        form = selectExperimentsForm(request.POST,initial={'prj_pk':prj_pk})
        if form.is_valid():
            if(slug=="EGA"):
                res = handle_exportegaform(request,prj_pk)
            elif(slug=="IHEC"):
                res = handle_exportiheacform(request,prj_pk)
            elif(slug=="GEO"):
                res = handle_exportgeoform(request,prj_pk)
                
            return res

    else:
        form = selectExperimentsForm(initial={'prj_pk':prj_pk})
     
    pageContext = {
        'form': form,
        'form_name':'Export '+slug+' form'
        }
    return render(request, 'export.html', pageContext)


